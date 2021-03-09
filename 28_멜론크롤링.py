import urllib.request as req
from bs4 import BeautifulSoup
import os # 이미지 다운 받는 모듈
import openpyxl # 엑셀 파일 모듈 #모듈 설치해야 할때 alt+enter+enter 또는 빨간전구
import datetime
import re
from openpyxl.drawing.image import Image

#이미지 저장할 폴더 생성
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

header = req.Request("https://www.melon.com/chart/index.htm", headers={"User-Agent":"Mozilla/5.0"})
# 위에 줄것을 안하면 멜론에서 해커인줄 알고 서버차단. 따라서 신원보장 입력
code = req.urlopen(header)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 > span > a") # 검사 눌러서 똑같은 경로있는지 검색 : ctrl + f
                                           # ellipsis rank02에서 띄어쓰기를 .으로 표현
name = soup.select("div.ellipsis.rank02 > span")
album = soup.select("div.ellipsis.rank03 > a")
img = soup.select("a.image_typeAll > img")
# for i in range(len(title)):
#     #print(title[i].string,name[i].string,album[i].string,img[i].attrs["src"])
#     # 위줄처럼 하면 두명의 가수일때 None으로 나온다. -> .string가 요소 내용을 추출하는데 문자열이 아닌 다른 요소들로 들어가서
#     # 이럴때는 name[i].string -> name[i].text 로 바꿔준다
#     print(title[i].string,name[i].text,album[i].string,img[i].attrs["src"])

#엑셀 파일 생성 자동화
if not os.path.exists("./멜론_크롤링.xlsx"):
    openpyxl.Workbook().save("./멜론_크롤링.xlsx") # 엑셀파일만드는 코드. 자주쓰임

book = openpyxl.load_workbook("./멜론_크롤링.xlsx") #엑셀 파일 불러와서 book에 저장
#엑셀 속 쓸데없는 시트는 삭세하기
if "Sheet" in book.sheetnames: # sheetnames를 통해 해당엑셀파일의 시트이름들을 리스트자료형으로 알려준다
    book.remove(book["Sheet"]) # 해당 리스트안에 한 원소가 있는지 알아보고 싶다면 in 사용
# sheet = book["sheet_name"] # 시트불러오기(1)
# sheet = book.active #시트불러오기(2)
sheet = book.create_sheet() # 시트이름을 오늘 날짜와 시간으로 만들기 1. 시트 생성
now = datetime.datetime.now() # 2.시간 불러오기
sheet.title = "{}년 {}월 {}일 {}분 {}초".format(now.year,now.month,now.day,now.minute,now.second)
row_num = 1
# 크롤링 결과 넣기 전에 행너비 지정
# 열 너비 조절
sheet.column_dimensions["A"].width = 15 # 해당 문장을 밑에 줄에 그대로 복사 :ctrl + d
sheet.column_dimensions["B"].width = 56
sheet.column_dimensions["C"].width = 29
sheet.column_dimensions["D"].width = 50

for i in range(len(title)):
    img_file_name = "./멜론이미지/{}.png".format(re.sub("[\\\/:*?\"<>\|]", " ", album[i].string))
    req.urlretrieve(img[i].attrs["src"], img_file_name)#(경로,"이미지 다운받을 파일명")
    #엑셀(시트)에 크롤링 결과 출력
    img_for_excel = Image(img_file_name)
    sheet.add_image(img_for_excel,"A{}".format(row_num))#엑셀에 이미지를 넣기 위한 코드가 따로 있다 #주의 :이미지를 넣을 칸의 이름넣기 : 알파벳과 숫자조합
    #sheet.cell(row=row_num, column=1).value =  #cell함수 : cell을 클릭하는 효과와 같다. 따라서 행열번호를 써주어야 한다.
    sheet.cell(row=row_num, column=2).value = title[i].string#.value를 넣지 않으면 그냥 셀을 클릭만 한거뿐, 셀 속에 내용 넣을 수 없다.
    sheet.cell(row=row_num, column=3).value = name[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    sheet.row_dimensions[row_num].height = 78 #행높이는 크롤링결과 기록할때마다 조절
    book.save("./멜론_크롤링.xlsx") # 갑자기 꺼지면 데이터 날라가니까 매행마다 저장
    print("{}위. {} - {}".format(row_num,title[i].string,name[i].text)) # row_num을 순위로 넣음
    row_num+=1