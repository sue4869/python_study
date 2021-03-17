#지역별로, 년도 별로 분리하고 차트 그리기

import pandas as pd
from pandas import ExcelWriter # 년도별로 시트 분리해야하는데 안되네? 이거 부르자
import openpyxl
from openpyxl.styles import Font, Alignment , PatternFill,Color, Border,Side
from openpyxl.chart import Reference, Series, BarChart

df = pd.read_excel("./엑셀데이터/아파트분양가격.xlsx")
#엑셀의 지역명이 어떤것이 있는지 확인
location_list = df["지역명"].unique() #unique() : 중복된 데이터를 빼주기 때문에 어떤 종류의 데이터가 있는지 확인 가능
#print(location_list) ---->['서울' '인천' '경기' '부산' '대구' '광주' '대전' '울산' '세종' '강원' '충북' '충남' '전북' '전남]

for location in location_list:
    df_location = df[df["지역명"] == location] #처음에 리스트가 서울이니 서울의 데이터만 가져온다.
    year_list = df_location["연도"].unique()#서울 것중에 어떤 년도가 있는지 뽑기
    #print(year_list) --> [2015 2016 2017 2018 2019 2020]
    writer = ExcelWriter("./아파트분양가격_{}.xlsx".format(location))
    for year in year_list:
        df_year = df_location[df_location["연도"] == year ]
        df_year = df_year.sort_values(by = "월", ascending=True) #월 기준으로 정렬
        df_year.to_excel(writer, sheet_name=str(year), index=None) #sheet_name에 문자열로 넣어져야 한다
    writer.save()                       # year을 정수형
    print("{}지역 엑셀 분리 완료".format(location))

#엑셀 서식
font = Font(name="맑은 고딕",size=12, bold=True)
alignmen_center = Alignment(horizontal="center")

color_pink = PatternFill(patternType="solid", fgColor=Color("FF848F"))
color_lightPink = PatternFill(patternType="solid", fgColor=Color("FFC6C3"))
color_lightPink2 = PatternFill(patternType="solid", fgColor=Color("FFE3EE"))

border = Border(left=Side(style="thin"),right=Side(style="thin"),
       top=Side(style="thin"),bottom=Side(style="thin"))

for location in location_list:
    book = openpyxl.load_workbook("./아파트분양가격_{}.xlsx".format(location)) #지역별 엑셀 파일 열기
    for sheet_name in book.sheetnames: #.sheetnames : 해당 엑셀파일의 시트 이름을 리스트 자료형으로 가져온다.
        sheet = book[sheet_name]
        sheet.column_dimensions["B"].width = 40 # 열 너비 조절
        #데이터 헤더 서식지정
        for row in sheet["A1:E1"]:
            for cell in row:
                cell.font = font
                cell.alignment = alignmen_center
                cell.fill = color_pink
                cell.border = border

        #데이터 내용들 서식 지정
        for row in sheet["A2:E{}".format(sheet.max_row)]: # 데이터 자료마다 끝 개수가 다르기에, 가장 끝행 번호
            for cell in row:
                cell.font = font
                cell.alignment = alignmen_center
                cell.fill = color_lightPink
                cell.border = border

        #바 차트 그리기
        chart = BarChart()
        chart.title = "{}지역 {}년도 아파트분양가격".format(location,year)
        val = Reference(sheet, range_string="{}!E2:E{}".format(sheet_name,sheet.max_row))
        series = Series(val)
        chart.append(series)
        sheet.add_chart(chart, "F1") #차트 위치 지정

    book.save("./아파트분양가격_{}.xlsx".format(location))
    print("{}지역 엑셀 파일 서식 지정 완료".format(location))

